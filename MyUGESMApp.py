import streamlit as st
import openpyxl
import os
import pywhatkit as kit  # Pour envoyer des messages WhatsApp
from openpyxl import load_workbook
import re  # Pour valider les numéros

# Chemin du fichier Excel
EXCEL_FILE = "etudiants.xlsx"

# Vérification et création du fichier Excel
if not os.path.exists(EXCEL_FILE):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Etudiants"
    sheet.append([
        "Prénom", "Nom", "Date de Naissance", "Lieu de Naissance",
        "Numéro WhatsApp Étudiant", "Adresse",
        "Nom Contact Maroc", "Prénom Contact Maroc", "Téléphone Contact Maroc", "Adresse Contact Maroc",
        "Nom Contact Sénégal", "Prénom Contact Sénégal", "Téléphone Contact Sénégal", "Adresse Contact Sénégal",
        "Établissement", "Filière", "Maladie Courante", "Groupe Sanguin"
    ])
    wb.save(EXCEL_FILE)

# Fonction pour sauvegarder les données dans Excel
def save_to_excel(data):
    wb = load_workbook(EXCEL_FILE)
    sheet = wb.active
    sheet.append(data)
    wb.save(EXCEL_FILE)

# Fonction pour valider les numéros
def validate_number(number):
    """
    Valide et nettoie le numéro de téléphone.
    - Accepte uniquement les numéros internationaux commençant par + suivi de chiffres.
    - Retire les espaces ou caractères inutiles.
    """
    number = number.replace(" ", "").replace("-", "").replace("(", "").replace(")", "")
    if re.match(r"^\+\d{10,15}$", number):
        return number  # Numéro valide
    else:
        return None  # Numéro invalide

# Fonction pour envoyer des messages WhatsApp
def send_whatsapp_messages(numbers, message):
    errors = []
    for number in numbers:
        try:
            kit.sendwhatmsg_instantly(number, message, wait_time=10, tab_close=True)
        except Exception as e:
            errors.append((number, str(e)))
    return errors

# Interface Streamlit
st.title("Recensement des Étudiants - UGESM")

# Mode utilisateur ou administrateur
mode = st.radio("Sélectionnez un mode :", ("Utilisateur", "Administrateur"))

if mode == "Utilisateur":
    st.subheader("Formulaire d'Enregistrement")
    with st.form("student_form"):
        st.write("Informations Générales")
        prenom = st.text_input("Prénom")
        nom = st.text_input("Nom")
        date_naissance = st.text_input("Date de Naissance (JJ/MM/AAAA)")
        lieu_naissance = st.text_input("Lieu de Naissance")
        numero_whatsapp = st.text_input("Numéro WhatsApp Étudiant")
        adresse = st.text_input("Adresse")
        etablissement = st.text_input("Établissement")
        filiere = st.text_input("Filière")
        maladie = st.text_input("Maladie Courante")
        groupe_sanguin = st.text_input("Groupe Sanguin")

        st.write("Contact d'Urgence au Maroc")
        nom_maroc = st.text_input("Nom Contact Maroc")
        prenom_maroc = st.text_input("Prénom Contact Maroc")
        telephone_maroc = st.text_input("Téléphone Contact Maroc")
        adresse_maroc = st.text_input("Adresse Contact Maroc")

        st.write("Contact d'Urgence au Sénégal")
        nom_senegal = st.text_input("Nom Contact Sénégal")
        prenom_senegal = st.text_input("Prénom Contact Sénégal")
        telephone_senegal = st.text_input("Téléphone Contact Sénégal")
        adresse_senegal = st.text_input("Adresse Contact Sénégal")

        submitted = st.form_submit_button("Soumettre")
        if submitted:
            if not all([prenom, nom, date_naissance, lieu_naissance, numero_whatsapp]):
                st.error("Veuillez remplir tous les champs obligatoires.")
            else:
                numero_whatsapp = validate_number(numero_whatsapp)
                if not numero_whatsapp:
                    st.error("Numéro WhatsApp invalide. Veuillez entrer un numéro au format international (+XXX...).")
                else:
                    student_data = [
                        prenom, nom, date_naissance, lieu_naissance, numero_whatsapp, adresse,
                        nom_maroc, prenom_maroc, telephone_maroc, adresse_maroc,
                        nom_senegal, prenom_senegal, telephone_senegal, adresse_senegal,
                        etablissement, filiere, maladie, groupe_sanguin
                    ]
                    save_to_excel(student_data)
                    st.success("Enregistrement réussi !")

elif mode == "Administrateur":
    st.subheader("Mode Administrateur")
    st.write("Accès aux données des étudiants et options d'envoi de messages.")

    # Chargement des données
    wb = load_workbook(EXCEL_FILE)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    headers = rows[0]
    data = rows[1:]

    # Affichage des données
    if len(data) == 0:
        st.warning("Aucun étudiant enregistré.")
    else:
        st.write(f"Nombre total d'étudiants enregistrés : {len(data)}")
        df = {headers[i]: [row[i] for row in data] for i in range(len(headers))}
        st.dataframe(df)

        # Envoi de messages groupés
        st.subheader("Envoi de Messages WhatsApp")
        message = st.text_area("Message à envoyer")
        if st.button("Envoyer Messages WhatsApp"):
            student_numbers = [row[4] for row in data if row[4]]
            valid_numbers = [validate_number(num) for num in student_numbers if validate_number(num)]
            invalid_numbers = [num for num in student_numbers if validate_number(num) is None]

            if not valid_numbers:
                st.warning("Aucun numéro valide trouvé.")
            elif not message:
                st.error("Veuillez entrer un message.")
            else:
                errors = send_whatsapp_messages(valid_numbers, message)
                if errors:
                    st.error(f"Échec de l'envoi pour certains numéros : {errors}")
                else:
                    st.success("Messages envoyés avec succès.")
            
            if invalid_numbers:
                st.warning(f"Numéros invalides détectés : {invalid_numbers}")
